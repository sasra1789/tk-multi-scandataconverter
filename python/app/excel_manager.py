
import os
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

def get_next_versioned_filename(base_path, prefix="scanlist", ext=".xlsx"):
    """
    기존 파일을 기준으로 다음 버전명을 자동 생성
    예: scanlist_v001.xlsx → scanlist_v002.xlsx
    """
    dir_name = os.path.dirname(base_path)
    base_name = os.path.splitext(os.path.basename(base_path))[0]

    # 파일명에서 prefix_v### 형식 추출
    pattern = re.compile(rf"{re.escape(prefix)}_v(\d{{3}})")
    existing_versions = []

    for f in os.listdir(dir_name):
        match = pattern.match(f)
        if match:
            existing_versions.append(int(match.group(1)))

    next_version = max(existing_versions, default=0) + 1
    return os.path.join(dir_name, f"{prefix}_v{next_version:03d}{ext}")

# def save_to_excel_with_thumbnails(data_list, save_path):
#     """
#     데이터를 엑셀(.xlsx) 파일로 저장하면서 썸네일 이미지도 삽입하고, 경로도 같이 기록
#     """
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Scan Data"

#     #  헤더 정의 (썸네일 경로 포함)
#     headers = ["Thumbnail", "Roll", "Shot Name", "Version", "Type", "Path", "Thumbnail Path"]
#     ws.append(headers)

#     for item in data_list:
#         # 텍스트 항목 먼저 삽입
#         row = [
#             "",  # Thumbnail 이미지 삽입 위치
#             item.get("roll", ""),
#             item.get("shot_name", ""),
#             item.get("version", ""),
#             item.get("type", ""),
#             item.get("path", ""),
#             item.get("thumbnail", ""),  #  썸네일 경로도 저장
#         ]
#         ws.append(row)

#         #  이미지 삽입
#         img_path = item.get("thumbnail", "")
#         if img_path and os.path.exists(img_path):
#             try:
#                 img = XLImage(img_path)
#                 img.width = 100
#                 img.height = 60
#                 cell = f"A{ws.max_row}"  # A열에 삽입
#                 ws.add_image(img, cell)
#                 ws.row_dimensions[ws.max_row].height = 45
#             except Exception as e:
#                 print(f" 이미지 삽입 실패: {img_path}\n{e}")

#     # 저장
#     wb.save(save_path)
#     print(f" 엑셀 저장 완료: {save_path}")

def save_to_excel_with_thumbnails(data_list, save_path):
    """
    데이터를 엑셀(.xlsx) 파일로 저장하면서 썸네일 이미지도 삽입하고, 경로도 같이 기록
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Scan Data"

    #  헤더 정의 (썸네일 경로 포함)
    headers = ["Thumbnail", "Roll", "Shot Name", "Version", "Type", "Path", "Thumbnail Path"]
    ws.append(headers)

    for item in data_list:
        # 텍스트 항목 먼저 삽입
        row = [
            "",  # Thumbnail 이미지 삽입 위치
            item.get("roll", ""),
            item.get("shot_name", ""),
            item.get("version", ""),
            item.get("type", ""),
            item.get("path", ""),
            item.get("thumbnail", ""),  # 썸네일 경로도 저장
        ]
        ws.append(row)

        #  이미지 삽입
        img_path = item.get("thumbnail", "")
        if img_path and os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                img.width = 100
                img.height = 60
                cell = f"A{ws.max_row}"  # A열에 삽입
                ws.add_image(img, cell)
                ws.row_dimensions[ws.max_row].height = 45
            except Exception as e:
                print(f"❌ 이미지 삽입 실패: {img_path}\n{e}")

    # 저장
    wb.save(save_path)
    print(f" 엑셀 저장 완료: {save_path}")

def load_excel_data(xlsx_path):
    """
    저장된 .xlsx 파일을 열어 딕셔너리 리스트로 반환
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data_list.append(row_data)

    return data_list



def get_next_versioned_filename(base_path):
    base_dir = os.path.dirname(base_path)
    base_name = os.path.splitext(os.path.basename(base_path))[0]  # scanlist
    ext = os.path.splitext(base_path)[1]  # .xlsx

    version = 1
    while True:
        filename = f"{base_name}_v{version:03d}{ext}"
        full_path = os.path.join(base_dir, filename)
        if not os.path.exists(full_path):
            return full_path
        version += 1
